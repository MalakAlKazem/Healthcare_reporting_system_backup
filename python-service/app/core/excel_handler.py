"""
Excel Handler - Parse mortality Excel files
Handles multi-sheet workbooks and extracts WHO categories
"""

import pandas as pd
import openpyxl
from typing import Dict, List, Tuple, Optional
from loguru import logger


class ExcelHandler:
    """Handle Excel file parsing"""
    
    def parse_excel(self, file_path: str) -> Tuple[pd.DataFrame, Optional[List[Dict]], Optional[Dict]]:
        """
        Parse Excel mortality file
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Tuple of (dataframe, who_summary, admission_data)
        """
        logger.info(f"📂 Parsing Excel: {file_path}")
        
        try:
            # Load workbook
            wb = openpyxl.load_workbook(file_path, data_only=True)
            logger.info(f"📚 Found {len(wb.sheetnames)} sheets: {', '.join(wb.sheetnames)}")
            
            # Parse main data (first sheet)
            main_sheet = wb.worksheets[0]
            df = self._parse_main_sheet(main_sheet)
            
            # Parse WHO categories (Sheet4 if exists)
            who_summary = self._parse_who_categories(wb)
            
            # Parse admission data (from 'الوفيات' sheet rows 44-45 if exists)
            admission_data = self._parse_admission_data(wb)
            
            logger.success(f"✅ Parsed {len(df)} records")
            
            return df, who_summary, admission_data
            
        except Exception as e:
            logger.error(f"❌ Excel parse error: {str(e)}")
            raise
    
    def _parse_main_sheet(self, sheet) -> pd.DataFrame:
        """Parse main mortality data sheet"""
        # Get all values
        data = []
        headers = []
        
        # Get headers from row 1
        for cell in sheet[1]:
            headers.append(str(cell.value).strip() if cell.value else '')
        
        # Get data rows
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if any(row):  # Skip empty rows
                data.append(row)
        
        # Create DataFrame
        df = pd.DataFrame(data, columns=headers)
        
        # Remove completely empty rows
        df = df.dropna(how='all')
        
        logger.debug(f"📊 Main sheet: {len(df)} rows × {len(df.columns)} columns")
        
        return df
    
    def _parse_who_categories(self, wb) -> Optional[List[Dict]]:
        """Parse WHO categories from Sheet4"""
        try:
            if 'Sheet4' not in wb.sheetnames:
                logger.debug("ℹ️  Sheet4 not found")
                return None
            
            sheet = wb['Sheet4']
            
            # Get header
            headers = [str(cell.value).strip() if cell.value else '' for cell in sheet[1]]
            
            # Find columns
            category_col = None
            count_col = None
            
            for idx, header in enumerate(headers):
                if 'تصنيف' in header or 'category' in header.lower():
                    category_col = idx
                if 'عدد' in header or 'count' in header.lower():
                    count_col = idx
            
            if category_col is None or count_col is None:
                logger.warning("⚠️  Could not find WHO category columns")
                return None
            
            # Parse data
            who_summary = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[category_col] and row[count_col]:
                    category = str(row[category_col]).strip()
                    count = row[count_col]
                    
                    if category and category != 'None':
                        try:
                            who_summary.append({
                                'category': category,
                                'count': int(count) if count else 0
                            })
                        except:
                            pass
            
            logger.success(f"✅ Parsed {len(who_summary)} WHO categories")
            return who_summary
            
        except Exception as e:
            logger.warning(f"⚠️  WHO categories parse failed: {str(e)}")
            return None
    
    def _parse_admission_data(self, wb) -> Optional[Dict]:
        """
        Parse admission data from 'الوفيات' sheet rows 44-45
        Row 44: BCI data
        Row 45: RAH data
        """
        try:
            sheet_name = 'الوفيات'
            if sheet_name not in wb.sheetnames:
                logger.debug(f"ℹ️  Sheet '{sheet_name}' not found")
                return None
            
            sheet = wb[sheet_name]
            
            # Row 44 = BCI
            row_44 = list(sheet[44])
            bci_building = str(row_44[4].value) if len(row_44) > 4 else 'BCI'
            bci_deaths = row_44[5].value if len(row_44) > 5 else 0
            bci_admissions = row_44[6].value if len(row_44) > 6 else 0
            
            # Row 45 = RAH
            row_45 = list(sheet[45])
            rah_building = str(row_45[4].value) if len(row_45) > 4 else 'RAH'
            rah_deaths = row_45[5].value if len(row_45) > 5 else 0
            rah_admissions = row_45[6].value if len(row_45) > 6 else 0
            
            # Convert to int
            bci_deaths = int(bci_deaths) if bci_deaths else 0
            bci_admissions = int(bci_admissions) if bci_admissions else 0
            rah_deaths = int(rah_deaths) if rah_deaths else 0
            rah_admissions = int(rah_admissions) if rah_admissions else 0
            
            admission_data = {
                'bci': {
                    'building': bci_building,
                    'deaths': bci_deaths,
                    'admissions': bci_admissions,
                    'rate': round((bci_deaths / bci_admissions * 100), 2) if bci_admissions > 0 else 0
                },
                'rah': {
                    'building': rah_building,
                    'deaths': rah_deaths,
                    'admissions': rah_admissions,
                    'rate': round((rah_deaths / rah_admissions * 100), 2) if rah_admissions > 0 else 0
                },
                'total': {
                    'deaths': bci_deaths + rah_deaths,
                    'admissions': bci_admissions + rah_admissions,
                    'rate': round(((bci_deaths + rah_deaths) / (bci_admissions + rah_admissions) * 100), 2) if (bci_admissions + rah_admissions) > 0 else 0
                }
            }
            
            logger.success(f"✅ Admission data: BCI={bci_deaths}/{bci_admissions}, RAH={rah_deaths}/{rah_admissions}")
            return admission_data
            
        except Exception as e:
            logger.warning(f"⚠️  Admission data parse failed: {str(e)}")
            return None


# Singleton
excel_handler = ExcelHandler()