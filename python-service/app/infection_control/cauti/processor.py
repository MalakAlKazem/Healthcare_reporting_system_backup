"""
CAUTI Excel Processor

Mirrors the VAP processor design:
  - Uses openpyxl (not pandas)
  - Normalises all column headers (strip + lowercase + collapse whitespace)
  - Returns a list of clean, normalized case dicts (one dict = one infection case)
  - Does NOT compute rates — that is done by InfectionControlStatistics

Usage:
    result = process_cauti_excel(file_path, year, quarter, denominators)
    cases  = result["cases"]
"""

import re
import math
import openpyxl
from datetime import datetime, date
from typing import Dict, List, Optional


# ── Column normalization map ──────────────────────────────────────────────────
COL_MAP: Dict[str, str] = {
    "year":                            "year",
    "semester":                        "semester",
    "type of infection":               "type_of_infection",
    "floor":                           "floor",
    "case number":                     "case_number",
    "nb of cases":                     "nb_of_cases",
    "age/year":                        "age",
    "diagnosis":                       "diagnosis",
    "date of admission":               "date_of_admission",
    "admission date":                  "date_of_admission",
    "date of foley insertion":         "date_of_insertion",
    "date of /intubation /insertion":  "date_of_insertion",
    "date of intubation/insertion":    "date_of_insertion",
    "date of intubation / insertion":  "date_of_insertion",
    "intubation date/ insertion":      "date_of_insertion",
    "intubation date/insertion":       "date_of_insertion",
    "تاريخ التنبيب":                   "date_of_insertion",
    "date of infection":               "date_of_infection",
    "infection date":                  "date_of_infection",
    "germs":                           "germs",
    "gender":                          "gender",
    "diabetic":                        "diabetic",
    "diabetes":                        "diabetic",
    "hypertension":                    "hypertension",
    "dyslipidemia":                    "dyslipidemia",
    "heart disease":                   "heart_disease",
    "kidney disease":                  "kidney_disease",
    "copd":                            "copd",
    "smoker":                          "smoker",
    "smoking":                         "smoker",
    "obesity":                         "obesity",
    "cardiac congenital malformation":  "cardiac_congenital_malformation",
    "&cardiac congenital malformation": "cardiac_congenital_malformation",
    "advanced age":                    "advanced_age",
    "length of stay":                  "length_of_stay",
    "duration of catheter":            "duration_of_catheter",
    "cancer":                          "cancer",
    "compromised immune system":       "compromised_immune_system",
    "respiratory pb":                  "respiratory_pb",
    "site of catheter":                "site_of_catheter_femoral",
    "site of catheter (femoral)":      "site_of_catheter_femoral",
    "femoral":                         "site_of_catheter_femoral",
    "femoral site":                    "site_of_catheter_femoral",
    "readmission date":                    "readmission_date",
    "discharge date":                      "discharge_date",
    "inserted by":                         "inserted_by",
    "location of insertion/intubation":    "location_of_insertion",
    "/removal catheter tube date":         "removal_date",
    "dwell time":                          "dwell_time",
    "prematurity":                         "prematurity",
    "neonates":                            "neonates",
    "infant":                              "infant",
    "total parenteral nutrition (tpn)":    "total_parenteral_nutrition",
    "consciousness":                       "consciousness",
    "head trauma":                         "head_trauma",
    "burns":                               "burns",
    "malnutrition":                        "malnutrition",
    "prolonged antibiotic exposure":       "prolonged_antibiotic_exposure",
    "reintubat/ion recatheterization":     "reintubation_recatheterization",
    "tracheostomy":                        "tracheostomy",
}

RISK_FACTOR_COLS = [
    "diabetic", "hypertension", "dyslipidemia", "heart_disease",
    "kidney_disease", "copd", "smoker", "obesity",
    "cardiac_congenital_malformation", "advanced_age",
    "length_of_stay", "duration_of_catheter",
    "cancer", "compromised_immune_system", "respiratory_pb",
    "site_of_catheter_femoral",
    "prematurity", "neonates", "infant", "total_parenteral_nutrition",
    "consciousness", "head_trauma", "burns", "malnutrition",
    "prolonged_antibiotic_exposure", "reintubation_recatheterization",
    "tracheostomy",
]


# ── Helpers ───────────────────────────────────────────────────────────────────

def _normalize_col(name) -> str:
    if name is None:
        return ""
    return " ".join(str(name).strip().lower().split())


def _safe(value):
    if value is None:
        return None
    try:
        if math.isnan(float(value)):
            return None
    except (TypeError, ValueError):
        pass
    return value


def _parse_age(value) -> Optional[float]:
    """Parse age to numeric years. Handles:
      • int / float               → taken as years directly
      • "53Y" / "3M" / "10D"     → single-letter suffix (Y=years, M=months, D=days)
      • "3 months" / "10 days"   → word form
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        try:
            if math.isnan(float(value)):
                return None
        except (TypeError, ValueError):
            return None
        return float(value)
    s = str(value).strip()
    sl = s.lower()
    m = re.match(r'^(\d+(?:\.\d+)?)\s*([ymd])$', sl)
    if m:
        num, unit = float(m.group(1)), m.group(2)
        if unit == 'y': return num
        if unit == 'm': return round(num / 12, 4)
        if unit == 'd': return round(num / 365, 4)
    if 'month' in sl:
        try:
            return round(float(sl.split()[0]) / 12, 4)
        except (ValueError, IndexError):
            return None
    if 'day' in sl:
        try:
            return round(float(sl.split()[0]) / 365, 4)
        except (ValueError, IndexError):
            return None
    if 'year' in sl:
        try:
            return float(sl.split()[0])
        except (ValueError, IndexError):
            return None
    try:
        return float(s)
    except ValueError:
        return None


def _age_display(age: Optional[float]) -> str:
    if age is None:
        return "N/A"
    if 0 < age < 1:
        months = round(age * 12)
        return f"{months}M" if months >= 1 else f"{round(age * 365)}D"
    return f"{int(age)}Y"


def _parse_date(value) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%m/%d/%Y"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    return None


def _days_between(d1, d2) -> Optional[int]:
    a, b = _parse_date(d1), _parse_date(d2)
    if a and b:
        return abs((b - a).days)
    return None


def _yes_no(value) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    if isinstance(value, str):
        return value.strip().lower() in ("yes", "oui", "1", "true", "نعم")
    return False


# ── Main processor ────────────────────────────────────────────────────────────

def process_cauti_excel(file_path: str, year: int, quarter: int, denominators: dict,
                        sheet_name=None) -> dict:
    """
    Read the CAUTI Excel file and return normalized case dicts.

    Returns
    -------
    {'cases': [normalized_case_dict, ...], 'denominators': denominators}
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)

    sheet = None
    if sheet_name is not None:
        sheet = wb[sheet_name]
    else:
        for name in wb.sheetnames:
            n = name.strip().upper()
            if "CAUTI" in n or n in ("SHEET1", "SHEET 1", "DATA", "CASES"):
                sheet = wb[name]
                break
        if sheet is None:
            sheet = wb.active

    raw_headers = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
    col_index: Dict[str, int] = {}
    for idx, raw in enumerate(raw_headers, start=1):
        norm   = _normalize_col(raw)
        mapped = COL_MAP.get(norm)
        if mapped:
            col_index[mapped] = idx

    if "case_number" not in col_index and "nb_of_cases" in col_index:
        col_index["case_number"] = col_index["nb_of_cases"]

    def get(row, field):
        idx = col_index.get(field)
        if idx is None:
            return None
        return _safe(sheet.cell(row, idx).value)

    cases: List[Dict] = []
    for r in range(2, sheet.max_row + 1):
        if all(sheet.cell(r, c).value is None for c in range(1, sheet.max_column + 1)):
            continue

        insertion    = _parse_date(get(r, "date_of_insertion"))
        infection    = _parse_date(get(r, "date_of_infection"))
        admission    = _parse_date(get(r, "date_of_admission"))
        readmission  = _parse_date(get(r, "readmission_date"))
        discharge    = _parse_date(get(r, "discharge_date"))
        removal      = _parse_date(get(r, "removal_date"))
        age          = _parse_age(get(r, "age"))

        case: Dict = {
            "case_number":          get(r, "case_number"),
            "nb_of_cases":          get(r, "nb_of_cases"),
            "year":                 get(r, "year") or year,
            "semester":             get(r, "semester"),
            "type_of_infection":    get(r, "type_of_infection"),
            "floor":                get(r, "floor"),
            "age":                  age,
            "age_display":          _age_display(age),
            "gender":               get(r, "gender"),
            "diagnosis":            get(r, "diagnosis"),
            "germs":                get(r, "germs"),
            "date_of_admission":    str(admission)   if admission   else None,
            "readmission_date":     str(readmission) if readmission else None,
            "discharge_date":       str(discharge)   if discharge   else None,
            "date_of_insertion":    str(insertion)   if insertion   else None,
            "removal_date":         str(removal)     if removal     else None,
            "date_of_infection":    str(infection)   if infection   else None,
            "infection_month":      infection.strftime("%Y-%m") if infection else None,
            "catheter_duration":    _days_between(insertion, infection),
            "dwell_time":           _safe(get(r, "dwell_time")),
            "inserted_by":          get(r, "inserted_by"),
            "location_of_insertion": get(r, "location_of_insertion"),
            **{col: _yes_no(get(r, col)) for col in RISK_FACTOR_COLS},
        }
        cases.append(case)

    return {"cases": cases, "denominators": denominators}
