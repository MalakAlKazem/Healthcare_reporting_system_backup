"""
VAP (Ventilator-Associated Pneumonia) Data Processor
Processes the VAP sheet from infection control Excel files.

FIXES applied:
  1. Added 'case number' to COL_MAP (was missing — used in tables but never mapped)
  2. Fixed leading-space column names: ' heart disease', ' kidney disease'
  3. Fixed _yes_no() to handle 'Yes'/'No' strings (Excel stores Yes/No not True/False)
  4. length_of_stay and duration_of_ventilation stored as Yes/No in this file — 
     treated as boolean presence flags; numeric value falls back to None gracefully
"""

import re
import math
import openpyxl
from datetime import datetime, date
from typing import Optional


# ─── Column name normalization map ──────────────────────────────────────────
COL_MAP = {
    # FIX 1: 'case number' was missing entirely
    "case number":                       "case_number",
    "year":                              "year",
    "semester":                          "semester",
    "type of infection":                 "type_of_infection",
    "floor":                             "floor",
    "nb of cases":                       "nb_of_cases",
    "age/year":                          "age",
    "diagnosis":                         "diagnosis",
    "date of admission":                 "date_of_admission",
    "date of intubation":                "date_of_intubation",
    "date of /intubation /insertion":    "date_of_intubation",
    "date of intubation/insertion":      "date_of_intubation",
    "date of intubation / insertion":    "date_of_intubation",
    "تاريخ التنبيب":                     "date_of_intubation",
    "date of infection":                 "date_of_infection",
    "germs":                             "germs",
    "gender":                            "gender",
    "diabetic":                          "diabetic",
    "hypertension":                      "hypertension",
    "dyslipidemia":                      "dyslipidemia",
    # FIX 2: Excel headers have a leading space — strip handles it in _normalize_col
    "heart disease":                     "heart_disease",
    "kidney disease":                    "kidney_disease",
    "copd":                              "copd",
    "smoker":                            "smoker",
    "obesity":                           "obesity",
    "cardiac congenital malformation":   "cardiac_congenital_malformation",
    "advanced age":                      "advanced_age",
    "length of stay":                    "length_of_stay",
    "duration of ventilation":           "duration_of_ventilation",
    "cancer":                            "cancer",
    "compromised immune system":         "compromised_immune_system",
    "respiratory pb":                    "respiratory_pb",
}

RISK_FACTOR_COLS = [
    "diabetic", "hypertension", "dyslipidemia", "heart_disease",
    "kidney_disease", "copd", "smoker", "obesity",
    "cardiac_congenital_malformation", "advanced_age",
    "cancer", "compromised_immune_system", "respiratory_pb",
]

RISK_FACTOR_LABELS = {
    "diabetic":                        "Diabetic / السكري",
    "hypertension":                    "Hypertension / ارتفاع ضغط الدم",
    "dyslipidemia":                    "Dyslipidemia / دسليبيدميا",
    "heart_disease":                   "Heart Disease / أمراض القلب",
    "kidney_disease":                  "Kidney Disease / أمراض الكلى",
    "copd":                            "COPD / مرض الانسداد الرئوي",
    "smoker":                          "Smoker / مدخن",
    "obesity":                         "Obesity / السمنة",
    "cardiac_congenital_malformation": "Cardiac Congenital Malformation / تشوه قلبي خلقي",
    "advanced_age":                    "Advanced Age / تقدم العمر",
    "cancer":                          "Cancer / السرطان",
    "compromised_immune_system":       "Compromised Immune System / ضعف المناعة",
    "respiratory_pb":                  "Respiratory Problem / مشكلة تنفسية",
}


# ─── Helpers ─────────────────────────────────────────────────────────────────

def _normalize_col(name: str) -> str:
    """Strip leading/trailing whitespace, lowercase, collapse internal spaces."""
    if name is None:
        return ""
    # FIX 2: strip() removes the leading space in ' Heart disease', ' kidney disease'
    return " ".join(str(name).strip().lower().split())


def _safe(value):
    """Return None for NaN, else the value."""
    if value is None:
        return None
    try:
        if math.isnan(float(value)):
            return None
    except (TypeError, ValueError):
        pass
    return value


def _to_float(value) -> Optional[float]:
    """Safely convert to float, return None if not numeric."""
    if value is None:
        return None
    try:
        v = float(value)
        if math.isnan(v):
            return None
        return v
    except (TypeError, ValueError):
        return None


def _age_display(age: Optional[float]) -> str:
    """Convert numeric age (years) to display string: '53Y', '3M', '10D'."""
    if age is None:
        return "N/A"
    if 0 < age < 1:
        months = round(age * 12)
        return f"{months}M" if months >= 1 else f"{round(age * 365)}D"
    return f"{int(age)}Y"


def _parse_age(value) -> Optional[float]:
    """
    Parse age to numeric years. Handles:
      • int / float                   → taken as years directly
      • "53Y" / "53y" / "53 Y"       → 53.0 years
      • "3M"  / "3m"  / "3 months"   → 0.25 years
      • "10D" / "10d" / "10 days"     → 0.027 years
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        try:
            if math.isnan(float(value)):
                return None
        except (TypeError, ValueError):
            return None
        return float(value)
    s = str(value).strip()
    sl = s.lower()
    # Single-letter suffix: 53Y / 3M / 10D
    m = re.match(r'^(\d+(?:\.\d+)?)\s*([ymd])$', sl)
    if m:
        num, unit = float(m.group(1)), m.group(2)
        if unit == 'y': return num
        if unit == 'm': return round(num / 12, 4)
        if unit == 'd': return round(num / 365, 4)
    # Word form: "3 months", "10 days"
    if 'month' in sl:
        try:
            return round(float(sl.split()[0]) / 12, 4)
        except (ValueError, IndexError):
            return None
    if 'day' in sl:
        try:
            return round(float(sl.split()[0]) / 365, 4)
        except (ValueError, IndexError):
            return None
    if 'year' in sl:
        try:
            return float(sl.split()[0])
        except (ValueError, IndexError):
            return None
    try:
        return float(s)
    except ValueError:
        return None


def _parse_date(value) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%m/%d/%Y"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    return None


def _days_between(d1, d2) -> Optional[int]:
    a = _parse_date(d1)
    b = _parse_date(d2)
    if a and b:
        return abs((b - a).days)
    return None


def _yes_no(value) -> bool:
    """
    FIX 3: Excel stores 'Yes'/'No' strings (not True/False booleans).
    Added 'yes'/'no' string handling.
    """
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    if isinstance(value, str):
        return value.strip().lower() in ("yes", "oui", "1", "true", "نعم")
    return False


def _avg(lst): return round(sum(lst) / len(lst), 1) if lst else None
def _min(lst): return min(lst) if lst else None
def _max(lst): return max(lst) if lst else None


# ─── Main processor ──────────────────────────────────────────────────────────

def process_vap_sheet(filepath: str,
                      ventilator_days: int,
                      semester: Optional[str] = None,
                      year: Optional[int] = None) -> dict:
    """
    Read the VAP sheet from an infection control Excel file and return
    a structured analytics dictionary.

    Parameters
    ----------
    filepath          : Path to the .xlsx file
    ventilator_days   : Total ventilator-days for the period (for overall VAP rate)
    semester          : Override semester label (e.g. 'First', 'Fourth')
    year              : Override year (int)

    Returns
    -------
    dict with keys:
        meta            – period info, total cases, VAP rate per 1,000 ventilator-days
        cases           – list of cleaned case records
        floors          – breakdown by floor/unit
        germs           – germ frequency (sorted desc)
        genders         – male/female/unknown counts
        age_groups      – age distribution in brackets
        risk_factors    – frequency + percentage for each comorbidity
        monthly_trend   – cases by infection month
        duration_stats  – avg/min/max for LOS, ventilation duration, onset delay
        diagnoses       – top admission diagnoses
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)

    sheet = None
    for name in wb.sheetnames:
        if name.strip().upper() == "VAP":
            sheet = wb[name]
            break
    if sheet is None:
        raise ValueError(f"No VAP sheet found. Available sheets: {wb.sheetnames}")

    # Build header → column index
    raw_headers = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
    col_index: dict[str, int] = {}
    for idx, raw in enumerate(raw_headers, start=1):
        norm = _normalize_col(raw)
        mapped = COL_MAP.get(norm)
        if mapped:
            col_index[mapped] = idx

    if "case_number" not in col_index and "nb_of_cases" in col_index:
        col_index["case_number"] = col_index["nb_of_cases"]

    def get(row, field):
        idx = col_index.get(field)
        if idx is None:
            return None
        return _safe(sheet.cell(row, idx).value)

    # Parse rows
    cases = []
    for r in range(2, sheet.max_row + 1):
        if all(sheet.cell(r, c).value is None for c in range(1, sheet.max_column + 1)):
            continue

        raw_admission  = get(r, "date_of_admission")
        raw_intubation = get(r, "date_of_intubation")
        raw_infection  = get(r, "date_of_infection")

        admission  = _parse_date(raw_admission)
        intubation = _parse_date(raw_intubation)
        infection  = _parse_date(raw_infection)

        # length_of_stay and duration_of_ventilation may be Yes/No flags in this Excel
        # _to_float returns None for 'Yes'/'No' strings which is correct behaviour
        los_raw = get(r, "length_of_stay")
        dov_raw = get(r, "duration_of_ventilation")

        age_val = _parse_age(get(r, "age"))
        case = {
            "case_number":                  get(r, "case_number"),
            "nb_of_cases":                  get(r, "nb_of_cases"),
            "year":                         get(r, "year"),
            "semester":                     get(r, "semester"),
            "floor":                        get(r, "floor"),
            "age":                          age_val,
            "age_display":                  _age_display(age_val),
            "gender":                       get(r, "gender"),
            "diagnosis":                    get(r, "diagnosis"),
            "germs":                        get(r, "germs"),
            "date_of_admission":            str(admission)  if admission  else None,
            "date_of_intubation":           str(intubation) if intubation else None,
            "date_of_infection":            str(infection)  if infection  else None,
            "infection_month":              infection.strftime("%Y-%m") if infection else None,
            "intubation_to_infection_days": _days_between(intubation, infection),
            "admission_to_infection_days":  _days_between(admission, infection),
            "length_of_stay":               _to_float(los_raw),
            "duration_of_ventilation":      _to_float(dov_raw),
            **{col: _yes_no(get(r, col)) for col in RISK_FACTOR_COLS},
        }
        cases.append(case)

    total_cases = len(cases)

    # VAP Rate per 1,000 ventilator-days
    vap_rate = round((total_cases / ventilator_days) * 1000, 2) if ventilator_days > 0 else None

    # Meta
    _year     = year     or (cases[0]["year"]     if cases else None)
    _semester = semester or (cases[0]["semester"] if cases else None)

    meta = {
        "year":                               _year,
        "semester":                           _semester,
        "total_cases":                        total_cases,
        "ventilator_days":                    ventilator_days,
        "vap_rate_per_1000_ventilator_days":  vap_rate,
    }

    # Floors
    floors: dict[str, int] = {}
    for c in cases:
        f = c["floor"] or "Unknown"
        floors[f] = floors.get(f, 0) + 1

    # Germs (sorted desc)
    germs: dict[str, int] = {}
    for c in cases:
        g = c["germs"] or "Unknown"
        germs[g] = germs.get(g, 0) + 1
    germs = dict(sorted(germs.items(), key=lambda x: x[1], reverse=True))

    # Genders
    genders: dict[str, int] = {"Male": 0, "Female": 0, "Unknown": 0}
    for c in cases:
        g = (c["gender"] or "").strip().capitalize()
        if g in ("Male", "Female"):
            genders[g] += 1
        else:
            genders["Unknown"] += 1

    # Age groups
    age_groups: dict[str, int] = {
        "<1": 0, "1-30": 0, "31-50": 0, "51-65": 0, "66-80": 0, "80+": 0, "Unknown": 0
    }
    for c in cases:
        a = c["age"]
        if a is None:
            age_groups["Unknown"] += 1
        elif a < 1:
            age_groups["<1"] += 1
        elif a <= 30:
            age_groups["1-30"] += 1
        elif a <= 50:
            age_groups["31-50"] += 1
        elif a <= 65:
            age_groups["51-65"] += 1
        elif a <= 80:
            age_groups["66-80"] += 1
        else:
            age_groups["80+"] += 1

    # Risk factors frequency
    risk_factors = []
    for col in RISK_FACTOR_COLS:
        count = sum(1 for c in cases if c.get(col))
        pct   = round(count / total_cases * 100, 1) if total_cases > 0 else 0
        risk_factors.append({
            "field":      col,
            "label":      RISK_FACTOR_LABELS[col],
            "count":      count,
            "percentage": pct,
        })
    risk_factors.sort(key=lambda x: x["count"], reverse=True)

    # Monthly trend
    monthly: dict[str, int] = {}
    for c in cases:
        m = c["infection_month"] or "Unknown"
        monthly[m] = monthly.get(m, 0) + 1
    monthly_trend = [{"month": k, "cases": v} for k, v in sorted(monthly.items())]

    # Duration statistics
    los_values   = [c["length_of_stay"]               for c in cases if c["length_of_stay"] is not None]
    vd_values    = [c["duration_of_ventilation"]       for c in cases if c["duration_of_ventilation"] is not None]
    onset_values = [c["intubation_to_infection_days"]  for c in cases if c["intubation_to_infection_days"] is not None]

    duration_stats = {
        "length_of_stay": {
            "avg": _avg(los_values), "min": _min(los_values), "max": _max(los_values),
            "note": "days from field 'length of stay'" if los_values else "no numeric data in file",
        },
        "duration_of_ventilation": {
            "avg": _avg(vd_values), "min": _min(vd_values), "max": _max(vd_values),
            "note": "days from field 'duration of ventilation'" if vd_values else "no numeric data in file",
        },
        "days_intubation_to_vap_onset": {
            "avg": _avg(onset_values), "min": _min(onset_values), "max": _max(onset_values),
            "note": "calculated: date_of_infection – date_of_intubation",
        },
    }

    # Top diagnoses
    diag_count: dict[str, int] = {}
    for c in cases:
        d = c["diagnosis"] or "Unknown"
        diag_count[d] = diag_count.get(d, 0) + 1
    diagnoses = [{"diagnosis": k, "count": v}
                 for k, v in sorted(diag_count.items(), key=lambda x: x[1], reverse=True)]

    return {
        "meta":           meta,
        "cases":          cases,
        "floors":         floors,
        "germs":          germs,
        "genders":        genders,
        "age_groups":     age_groups,
        "risk_factors":   risk_factors,
        "monthly_trend":  monthly_trend,
        "duration_stats": duration_stats,
        "diagnoses":      diagnoses,
    }